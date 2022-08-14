import os
from openpyxl import Workbook
from random import *

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = Workbook() # 새 워크북 생성
ws = wb.active

ws.title = 'Sheet'

ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3

ws['B1'] = 4
ws['B2'] = 5
ws['B3'] = 6

print(ws['A1']) # A1 셀의 정보를 출력
print(ws['A1'].value) # A1 셀의 값을 출력

print(ws['A10'].value) # 값이 없을땐 None 출력

# row = 1, 2, 3, ...
# column = A, B, C
print(ws.cell(row=1, column=1).value) # ws['A1'].value
print(ws.cell(row=1, column=2).value) # ws['B1'].value

c = ws.cell(row=1, column=3, value=10)
print(c.value) # ws['C1'].value

# 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for i in range(1, 11):
    for j in range(1, 11):
        # ws.cell(row=i, column=j, value=randint(0, 100))
        ws.cell(row=i, column=j, value=index)
        index += 1

wb.save('./out/sample.xlsx')

wb.close()