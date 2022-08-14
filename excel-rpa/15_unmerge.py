import os
from openpyxl import load_workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)

wb = load_workbook('./out/sample_merge.xlsx')
ws = wb.active

# B2:D2 병합되어 있던 셀을 해제
ws.unmerge_cells('B2:D2')
wb.save('./out/sample_unmerge.xlsx')
