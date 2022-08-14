import os, datetime
from openpyxl import Workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)

wb = Workbook()
ws = wb.active

ws['A1'] = datetime.datetime.today()
ws['A2'] = '=SUM(1, 2, 3)'
ws['A3'] = '=AVERAGE(1, 2, 3)'

ws['A4'] = 10
ws['A5'] = 20
ws['A6'] = '=SUM(A4:A5)'

wb.save('./out/sample_formular.xlsx')
