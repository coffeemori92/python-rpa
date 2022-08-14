import os
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = load_workbook('./out/sample.xlsx') # 파일에서 워크북 불러옴
ws = wb.active

a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']

# A열의 너비를 5로 설정
ws.column_dimensions['A'].width = 5

# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color='ff0000', italic=True, bold=True)
b1.font = Font(color='cc33ff', name='Arial', strike=True)
c1.font = Font(color='0000ff', size=20, underline='single')

# 테두리 적용
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 셀에 대해서 정렬
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # A 번호열은 제외
        if cell.column == 1: 
            continue
        # cell이 정수형 데이터 이고 90점 보다 크면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor='00ff00', fill_type='solid')
            cell.font = Font(color='ff0000')

# 틀 고정
ws.freeze_panes = 'B2'

wb.save('./out/sample_style.xlsx')
