import os
from openpyxl import load_workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
# wb = load_workbook('./out/sample_formular.xlsx') # 파일에서 워크북 불러옴
# ws = wb.active

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)
        

wb = load_workbook('./out/sample_formular.xlsx', data_only=True) # 파일에서 워크북 불러옴
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None이라고 표시
for row in ws.values:
    for cell in row:
        print(cell)
        