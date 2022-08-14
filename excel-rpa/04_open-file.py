import os
from openpyxl import load_workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = load_workbook('./out/sample.xlsx') # 파일에서 워크북 불러옴
ws = wb.active

# cell 데이터 불러오기
# for i in range(1, 11):
#     for j in range(1, 11):
#         print(ws.cell(row=i, column=j).value, end=' ')
#     print()

# cell 갯수를 모를 때
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        print(ws.cell(row=i, column=j).value, end=' ')
    print()