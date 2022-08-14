import os
from openpyxl import Workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = Workbook() # 새 워크북 생성

ws = wb.create_sheet() # 새로운 시트 기본이름으로 생성
ws.title = 'MySheet'   # 시트 이름 변경
ws.sheet_properties.tabColor = 'ff66ff' # RGB 형태로 값을 넣어, 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet('YourSheet')
# Sheet, MySheet, NewSheet, YourSheet
ws2 = wb.create_sheet('NewSheet', 2) # 2번째 index에 시트 생성

new_ws = wb['NewSheet'] # Dict 형태로 시트에 접급

# 시트 복사
new_ws['A1'] = 'Test'
target = wb.copy_worksheet(new_ws)
target.title = 'CopiedSheet'

print(wb.sheetnames) # 모든 Sheet 이름 확인

wb.save('./out/sample.xlsx')

wb.close()