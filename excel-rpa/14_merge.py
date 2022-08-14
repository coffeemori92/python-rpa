import os
from openpyxl import Workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)

wb = Workbook()
ws = wb.active

# B2부터 D2까지 셀을 합치겠음
ws.merge_cells('B2:D2')
ws['B2'].value = 'Merged Cell'

wb.save('./out/sample_merge.xlsx')
