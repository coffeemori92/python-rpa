import os
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import *

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = Workbook() # 새 워크북 생성
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(['번호', '영어', '수학'])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(1, 100)])

# col_B = ws['B'] # 영어 column만 가져오기
# print(col_B)

# for cell in col_B:
#     print(cell.value)
    
# col_range = ws['B:C'] # 영어, 수학 column 가져오기
# print(col_range)
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)
        
# row_title = ws[1] # 1번째 row 가져오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째 row에서 6번째 row 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=' ')
#     print()

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=' ')
#         # print(cell.coordinate, end=' ')
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=' ')
#         print(xy[0], end='')  # A
#         print(xy[1], end=' ') # 1
#     print()

# 전체 rows
# print(tuple(ws.rows))

# for row in tuple(ws.rows):
#     print(row[1].value)

# 전체 columns
# print(tuple(ws.columns))

# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 rows
#     print(row[1].value)
    
# for column in ws.iter_cols(): # 전체 columns
#     print(column[0].value)

# 1번째 row부터 5번째 row까지, 2번째 column부터 3번째 column까지
# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
#     print(row)
    
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save('./out/sample.xlsx')
