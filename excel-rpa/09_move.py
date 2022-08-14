import os
from openpyxl import load_workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = load_workbook('./out/sample.xlsx') # 파일에서 워크북 불러옴
ws = wb.active

# ws.move_range('B1:C11', rows=0, cols=1)
# ws['B1'].value = '국어'

ws.move_range('C1:C11', rows=5, cols=-1)

wb.save('./out/sample_korean.xlsx')
