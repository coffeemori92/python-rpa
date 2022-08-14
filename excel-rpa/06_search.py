import os
from openpyxl import load_workbook

try:
    if not os.path.exists('out'):
        os.mkdir('out')
except Exception as e:
    print(e)
    
wb = load_workbook('./out/sample.xlsx') # 파일에서 워크북 불러옴
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 번호, 영어 수학
    if int(row[1].value) > 80:
        print(row[0].value, '번 학생은 영어가 80점 초과')
        
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == '영어':
            cell.value = '컴퓨터'
            
wb.save('./out/sample_modified.xlsx')
